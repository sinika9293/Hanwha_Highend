"""schema.py 구조의 4개 섹션 데이터를 markdown 리포트와 Excel 표로 렌더링한다.

이 모듈은 API를 호출하지 않는다 — collect.py 가 만든
{"trend": ..., "amenity": ..., "facility": ..., "global_case": ...} dict를 받아
통합 리포트 파일로 바꾸는 순수 렌더링 담당이다.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from dates import ReportPeriod


def _bullet_list(items: list[str], numbered: bool = True) -> str:
    if numbered:
        return "\n".join(f"{i + 1}. {item}" for i, item in enumerate(items))
    return "\n".join(f"- {item}" for item in items)


def _subscription_table_md(rows: list[dict]) -> str:
    header = "| 구분 | 모집 | 접수 | 1순위 평균 경쟁률 |\n|---|---|---|---|"
    lines = [header]
    for row in rows:
        lines.append(
            f"| {row.get('label', '')} | {row.get('listings', '')} | "
            f"{row.get('applicants', '')} | **{row.get('competition_rate', '')}** |"
        )
    return "\n".join(lines)


def _deals_table_md(rows: list[dict]) -> str:
    header = "| 프로젝트 | 지역 | 뉴스 |\n|---|---|---|"
    lines = [header]
    for row in rows:
        lines.append(f"| {row.get('project', '')} | {row.get('region', '')} | {row.get('news', '')} |")
    return "\n".join(lines)


def _keywords_table_md(rows: list[dict]) -> str:
    header = "| 순위 | 키워드 | 왜 지금 | 확인 지표 |\n|---|---|---|---|"
    lines = [header]
    for row in sorted(rows, key=lambda r: r.get("rank", 0)):
        lines.append(
            f"| {row.get('rank', '')} | **{row.get('keyword', '')}** | "
            f"{row.get('why_now', '')} | {row.get('indicator', '')} |"
        )
    return "\n".join(lines)


def _items_md(items: list[dict]) -> str:
    lines = []
    for item in items:
        lines.append(f"- **{item.get('name', '')}**: {item.get('description', '')}")
    return "\n".join(lines)


def render_markdown(sections: dict, period: ReportPeriod, generated_on: date | None = None) -> str:
    if generated_on is None:
        generated_on = date.today()

    trend = sections.get("trend", {})
    amenity = sections.get("amenity", {})
    facility = sections.get("facility", {})
    global_case = sections.get("global_case", {})

    parts = [
        f"# 하이엔드 주거 트렌드 리포트 — {period.year}년 {period.month}월",
        "",
        f"작성일: {generated_on:%Y-%m-%d} | 대상 기간: {period.range_label}",
        "수집 채널: 국내 부동산·건설 매체, 해외 리서치하우스·건축 전문지, 브랜디드 레지던스 전문 매체",
        "",
        "---",
        "",
        "## 0. 한눈에 보기 (Executive Summary)",
        "",
        _bullet_list(trend.get("executive_summary", [])),
        "",
        "---",
        "",
        "## 1. 국내 동향",
        "",
        "### 1-1. 청약 양극화: 숫자로 본 격차",
        "",
        _subscription_table_md(trend.get("subscription_table", [])),
        "",
        trend.get("subscription_note", ""),
        "",
        "### 1-2. 이달 청약 결산",
        "",
        trend.get("monthly_summary", ""),
        "",
        "### 1-3. 정책·세제 변수",
        "",
        trend.get("policy_tax", ""),
        "",
        "### 1-4. 리스크 신호",
        "",
        _bullet_list(trend.get("risk_signals", []), numbered=False),
        "",
        "---",
        "",
        "## 2. 최신 어메니티 트렌드 (서비스·운영)",
        "",
        amenity.get("intro", ""),
        "",
        _items_md(amenity.get("items", [])),
        "",
        "---",
        "",
        "## 3. 떠오르는 부대시설 (하드웨어 공간)",
        "",
        facility.get("intro", ""),
        "",
        _items_md(facility.get("items", [])),
        "",
        "---",
        "",
        "## 4. 해외 개발 사례",
        "",
        "### 4-1. 시장 개요",
        "",
        global_case.get("market_overview", ""),
        "",
        "### 4-2. 이달 주요 딜",
        "",
        _deals_table_md(global_case.get("deals", [])),
        "",
        "### 4-3. 건축·설계 트렌드",
        "",
        global_case.get("design_trends", ""),
        "",
        "---",
        "",
        "## 5. 키워드 / 관전 포인트",
        "",
        "### 5-1. 요즘 관심도가 높은 키워드",
        "",
        _keywords_table_md(trend.get("keywords", [])),
        "",
        "### 5-2. 관전 포인트",
        "",
        _bullet_list(trend.get("watchpoints", [])),
        "",
        "---",
        "",
        "## 6. 출처",
        "",
        "**국내**",
        _bullet_list(trend.get("sources", []), numbered=False),
        "",
        "**해외**",
        _bullet_list(global_case.get("sources", []), numbered=False),
        "",
        "---",
        "",
        "*본 리포트는 Claude API 웹검색으로 수집한 공개 보도·리서치 자료를 기반으로 "
        "자동 생성했습니다. 수치·사실은 원 보도 시점 기준이며 별도 검증이 필요합니다.*",
        "",
    ]
    return "\n".join(parts)


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 60)


def _write_table(ws, headers: list[str], rows: list[list[str]]) -> None:
    bold = Font(bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    _autosize(ws)


def render_excel(sections: dict, period: ReportPeriod, path: Path) -> None:
    trend = sections.get("trend", {})
    amenity = sections.get("amenity", {})
    facility = sections.get("facility", {})
    global_case = sections.get("global_case", {})

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "청약경쟁률"
    _write_table(
        ws1,
        ["구분", "모집", "접수", "1순위 평균 경쟁률"],
        [
            [r.get("label", ""), r.get("listings", ""), r.get("applicants", ""), r.get("competition_rate", "")]
            for r in trend.get("subscription_table", [])
        ],
    )

    ws2 = wb.create_sheet("어메니티")
    _write_table(
        ws2,
        ["이름", "설명"],
        [[i.get("name", ""), i.get("description", "")] for i in amenity.get("items", [])],
    )

    ws3 = wb.create_sheet("부대시설")
    _write_table(
        ws3,
        ["이름", "설명"],
        [[i.get("name", ""), i.get("description", "")] for i in facility.get("items", [])],
    )

    ws4 = wb.create_sheet("해외딜")
    _write_table(
        ws4,
        ["프로젝트", "지역", "뉴스"],
        [
            [d.get("project", ""), d.get("region", ""), d.get("news", "")]
            for d in global_case.get("deals", [])
        ],
    )

    ws5 = wb.create_sheet("키워드")
    _write_table(
        ws5,
        ["순위", "키워드", "왜 지금", "확인 지표"],
        [
            [k.get("rank", ""), k.get("keyword", ""), k.get("why_now", ""), k.get("indicator", "")]
            for k in sorted(trend.get("keywords", []), key=lambda r: r.get("rank", 0))
        ],
    )

    ws6 = wb.create_sheet("메타")
    ws6.append(["대상 기간", period.range_label])
    ws6.append(["생성 파일", period.label])
    _autosize(ws6)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
